import os.path
from multiprocessing.dummy import Pool as ThreadPool, Lock
from multiprocessing import Process
import csv

import multiprocessing.dummy
import re

import sys
import time
from dateutil import parser as dateparser
from openpyxl import load_workbook

from datetime import datetime
import settings
from sqlalchemy import *
from sqlalchemy import event, exc
from sqlalchemy.orm import sessionmaker, scoped_session
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import relationship
import warnings
import os
import sqlalchemy.exc
from pprint import pprint
import tweet_api
import random

IS_PROFILE_SEARCH = False
ISUSERPROFILE = True
time_wait = 0
flag1 = False

emoticons_str = r"""
    (?:
        [:=;] # Eyes
        [oO\-]? # Nose (optional)
        [D\)\]\(\]/\\OpP] # Mouth
    )"""
regex_str = [
    # emoticons_str,
    r'<[^>]+>',  # HTML tags
    r'(?:@[\w_]+)',  # @-mentions
    r"(?:\#\w*[a-zA-Z]+\w*)",    # hash-tags
    r"(?:\$[A-Za-z0-9]{1,5}\b)",    # cash-tags
    r'http[s]?://(?:[a-z]|[0-9]|[$-_@.&amp;+]|[!*\(\),]|(?:%[0-9a-f][0-9a-f]))+',  # URLs

    r'(?:(?:\d+,?)+(?:\.?\d+)?)',  # numbers
    r"(?:[a-z][a-z'\-_]+[a-z])",  # words with - and '
    r'(?:[\w_]+)',  # other words
    r'(?:\S)'  # anything else
]
tokens_re = re.compile(r'(' + '|'.join(regex_str) + ')', re.VERBOSE | re.IGNORECASE)
emoticon_re = re.compile(r'^' + emoticons_str + '$', re.VERBOSE | re.IGNORECASE)


class Twit:
    def __init__(self):
        pass

    def json(self):
        return {'date': self.unixtime, 'text': self.text, 'screen_name': self.screen_name, 'user_name': self.user_name,
                'user_id': self.user_id,
                'id': self.id, 'retweets_count': self.retweets_count, 'favorites_count': self.favorites_count,
                'permalink': self.permalink, 'urls': self.urls, 'mentions': self.mentions,
                'hashtags': self.hashtags, 'is_retweet': self.is_retweet, 'symbols': self.symbols,
                'is_protected': self.is_protected}

    def __repr__(self):
        return self.text


def get_symbols(s):
    s = s.upper()
    res = re.findall('(\$[A-Z]{1,6}([._][A-Z]{1,2})?)', s, re.M)
    if res:
        r = list(map(lambda x: x[0], res))
        r = list(set(r))
        return r
    else:
        return []


def scra(query, i, proxy, lock, session):
    twitter_scraper = tweet_api.TweetScraper(proxy, IS_PROFILE_SEARCH=IS_PROFILE_SEARCH, logname='awam')

    def tokenize(s):
        return tokens_re.findall(s)

    def preprocess(s, lowercase=False):
        tokens = tokenize(s)
        if lowercase:
            tokens = [token if emoticon_re.search(token) else token.lower() for token in tokens]
        return tokens

    permno = query[4]
    fieldnames = ["QueryStartDate", "QueryEndDate", "Query", "DateOfActivity", "UserScreenName", "Keyword",
                  "Location", "Website", "DateJoined", "IsMention", "UserID", "TimeOfActivity", "Hashtags",
                  "Re_tweet", "NumberOfReplies", "NumberOfRe_tweets", "NumberOfFavorites", "Tweet", "tweet_id",
                  "tweet_url", "is_verified", "Urls", "UserFollowersCount", "UserFollowingCount", "UserTweetsCount",
                  "LikesCount", "CashtagSymbols", "user_location", "permno"]

    count = 0

    # ttm = time.time()
    # tweet_list = []

    # Receiving random credentials from the settings file
    credential = settings.twitter_logins[random.randrange(0, len(settings.twitter_logins))]
    login = credential['login']
    password = credential['password']

    for t in twitter_scraper.get_new_search(query, login, password):
        data = {}
        data['tweet_id'] = t.id
        print(query, count, t.date)
        twit = session.query(Tweet).filter_by(tweet_id=data['tweet_id']).first()
        if twit and ISUSERPROFILE:
            continue
        data['permno'] = permno
        data['user_location'] = t.user_location
        data['LikesCount'] = t.likes
        data['Website'] = t.website
        data['QueryStartDate'] = query[2]
        data['QueryEndDate'] = query[3]
        data['Query'] = query[1]
        data['Keyword'] = query[1]
        data['TimeOfActivity'] = time.strftime('%H:%M:%S', time.localtime(t.unixtime))
        data['DateOfActivity'] = time.strftime('%d/%m/%Y', time.localtime(t.unixtime))
        data['tweet_url'] = t.permalink
        data['UserID'] = t.user_id
        data['UserScreenName'] = t.screen_name
        data['UserName'] = t.user_name
        data['TimeZone'] = t.time_zone
        data['UserTweetsCount'] = t.user_tweet_count
        data['UserFollowersCount'] = t.user_followers_count
        data['UserFollowingCount'] = t.user_following_count
        data['NumberOfFavorites'] = t.favorites_count
        data['NumberOfRe_tweets'] = t.retweets_count
        data['NumberOfReplies'] = t.replyes_count
        data['Re_tweet'] = t.is_retweet
        data['is_verified'] = t.is_verified
        # data['isProtected'] = t.is_protected
        data['isReply'] = t.is_reply
        try:
            data['ReplyTweetID'] = int(t.data_conversation_id)
        except ValueError:
            data['ReplyTweetID'] = None
        data['ReplyUserId'] = t.is_reply_id
        # data['ReplyScreenName'] = t.is_reply_screen_name
        # data['Lang'] = t.lang
        data['Hashtags'] = ' '.join(t.hashtags)
        data['Urls'] = ', '.join(t.urls)
        data['CashtagSymbols'] = '\n'.join(t.symbols)
        # data['Mentions_id'] = '\n'.join(t.mentions_id)
        data['IsMention'] = ' '.join(t.mentions_name)
        data['Location'] = t.location_name
        # data['Location ID'] = t.location_id
        data['DateJoined'] = dateparser.parse(t.user_created) if t.user_created else None
        data['Tweet'] = t.text

        if hasattr(t, 'utc_offset') and t.utc_offset:
            if t.utc_offset / 3600 >= 0:
                data['TimeZoneUTC'] = 'UTC+' + str(int(t.utc_offset / 3600))
            else:
                data['TimeZoneUTC'] = 'UTC' + str(int(t.utc_offset / 3600))
        else:
            data['TimeZoneUTC'] = None

        tokens = preprocess(t.text)
        cashtags = set([term.upper() for term in tokens if term.startswith('$') and len(term) > 1])
        hashtags = set([term for term in tokens if term.startswith('#') and len(term) > 1])
        # mentions = [term for term in tokens if term.startswith('@') and len(term) > 1]
        urls = set([term for term in tokens if term.startswith('http') and len(term) > 4])

        # tweet_list.append(data)
        # pprint(data)

        # print(query, count, t.date)
        # print(data)
        if ISUSERPROFILE:
            pass
            # dd = re.findall('\w+ (\w+) (\d+) \d+:\d+:\d+ \+\d+ (\d+)', data['DateJoined'])
            # try:
            #     date_joined = datetime.strptime(' '.join(dd[0]), '%b %d %Y')
            # except IndexError:
            #     print(data['DateJoined'], dd)
            #     exit()
        else:
            date_joined = None
            t.user_listed_count = None
            t.username = t.user_name
            t.user_timezone = None
            t.user_description = None
            data['Website'] = None
            data['is_verified'] = None

        # if session.query(Tweet).filter_by(tweet_id=data['tweet_id']).first():
        #     continue

        user = session.query(User).filter_by(user_id=data['UserID']).first()
        if not user:
            user = User(user_id=data['UserID'],
                        twitter_handle=data['UserScreenName'][:120],
                        user_name=data['UserName'][:120],
                        location=data['user_location'][:255] if data['user_location'] else None,
                        date_joined=data['DateJoined'],
                        timezone=data['TimeZoneUTC'][:10] if data['TimeZoneUTC'] else None,
                        website=data['Website'][:255] if data['Website'] else None,
                        user_intro=t.user_description[:255] if t.user_description else None,
                        verified=data['is_verified'])
            if not session.query(UserCount).filter_by(user_id=data['UserID']).first():
                user_count = UserCount(follower=data['UserFollowersCount'],
                                       following=data['UserFollowingCount'],
                                       tweets=data['UserTweetsCount'],
                                       likes=data['NumberOfFavorites'],
                                       lists=t.user_listed_count)
            try:
                session.add(user_count)
                user.counts.append(user_count)
                session.add(user)
                session.commit()
                print('Inserted new user:', data['UserID'])
            except sqlalchemy.exc.IntegrityError as err:
                if re.search("duplicate key value violates unique constraint", err.args[0]):
                    print('ROLLBACK USER')
                    session.rollback()
            except Exception as e:
                print(e)
                raise

        # twit = session.query(Tweet).filter_by(tweet_id=data['tweet_id']).first()
        # if not twit:
        twit = Tweet(tweet_id=data['tweet_id'],
                     date=datetime.strptime(data['DateOfActivity'], '%d/%m/%Y'),
                     time=data['TimeOfActivity'],
                     timezone=data['TimeZoneUTC'][:10] if data['TimeZoneUTC'] else None,
                     retweet_status=data['Re_tweet'],
                     text=data['Tweet'],
                     reply_to=data['ReplyTweetID'],
                     location=data['Location'][:255] if data['Location'] else None,
                     permalink=data['tweet_url'][:255] if data['tweet_url'] else None,
                     emoticon=','.join(t.emoji) if t.emoji else None)
        tweet_count = TweetCount(reply=data['NumberOfReplies'],
                                 favorite=data['NumberOfFavorites'],
                                 retweet=t.retweets_count)

        if not session.query(TweetHashtags).filter_by(tweet_id=data['tweet_id']).first():
            for hash_s in hashtags:
                tweet_hashtag = TweetHashtags(hashtags=hash_s[:45])
                twit.hash_s.append(tweet_hashtag)
                session.add(tweet_hashtag)

        if not session.query(TweetUrl).filter_by(tweet_id=data['tweet_id']).first():
            for url_s in urls:
                tweet_url = TweetUrl(url=url_s[:255])
                twit.url_s.append(tweet_url)
                session.add(tweet_url)

        if not session.query(TweetCashtags).filter_by(tweet_id=data['tweet_id']).first():
            for cash_s in cashtags:
                tweet_cashtag = TweetCashtags(cashtags=cash_s[:45])
                twit.cash_s.append(tweet_cashtag)
                session.add(tweet_cashtag)

        if not session.query(TweetMentions).filter_by(tweet_id=data['tweet_id']).first():
            for ment_s in t.ment_s:
                tweet_mentions = TweetMentions(mentions=ment_s[0][:45], user_id=ment_s[1])
                twit.ment_s.append(tweet_mentions)
                session.add(tweet_mentions)
        user.tweets.append(twit)

        if not session.query(TweetCount).filter_by(tweet_id=int(data['tweet_id'])).first():
            twit.counts.append(tweet_count)
            session.add(tweet_count)
        else:
            i = 1
        try:
            session.add(twit)
            session.commit()
            print('Inserted new Tweet:', data['tweet_id'])
        except sqlalchemy.exc.IntegrityError as err:
            if re.search('duplicate key value violates unique constraint', err.args[0]):
                print('ROLLBACK common')
                session.rollback()
        except Exception as e:
            print(e)
            raise
        count += 1

    lock.acquire()
    if count > 0:

        with open('report.csv', 'a') as f:
            data = {}
            fdnames = ['time', 'query_name', 'number', ]
            writer = csv.DictWriter(f, lineterminator='\n', fieldnames=fdnames, dialect='excel', quotechar='"',
                                    quoting=csv.QUOTE_ALL)
            data['time'] = time.strftime('%Y-%m-%d %H:%M:%S')
            data['query_name'] = query[1]
            data['number'] = count
            # data['date'] = t.user_created

            writer.writerow(data)
        lock.release()
        return count
    else:
        with open('error.csv', 'a') as f:
            data = {}
            fdnames = ['time', 'query_name', 'number', ]
            writer = csv.DictWriter(f, lineterminator='\n', fieldnames=fdnames, dialect='excel', quotechar='"',
                                    quoting=csv.QUOTE_ALL)
            data['time'] = time.strftime('%Y-%m-%d %H:%M:%S')
            data['query_name'] = query[1]
            data['number'] = count
            writer.writerow(data)
        lock.release()
        return False


def scrape_query(user_queue, proxy, lock, pg_dsn):
    # db_engine = create_engine(pg_dsn, pool_size=1)
    add_engine_pidguard(db_engine)
    # DstSession = sessionmaker(bind=db_engine, autoflush=False)
    session = DstSession()

    active = True
    while not user_queue.empty():
        query, i = user_queue.get()

        print('START', i, proxy, query)
        # TODO filter:nativeretweets
        try:
            res = scra(query, i, proxy, lock, session=session)
        except tweet_api.LoadingError:
            print('LoadingError except')
            return False
        if not res:
            print('     SCRAP_USER Error in', query, i)
            with open('error_list.txt', 'a') as f:
                f.write(query[0] + '\n')
        else:
            print(datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ' ENDED', i, proxy, query, res)
    session.close()
    return True


def add_engine_pidguard(engine):
    @event.listens_for(engine, "connect")
    def connect(dbapi_connection, connection_record):
        connection_record.info['pid'] = os.getpid()

    @event.listens_for(engine, "checkout")
    def checkout(dbapi_connection, connection_record, connection_proxy):
        pid = os.getpid()
        if connection_record.info['pid'] != pid:
            # substitute log.debug() or similar here as desired
            warnings.warn(
                "Parent process %(orig)s forked (%(newproc)s) with an open "
                "database connection, "
                "which is being discarded and recreated." %
                {"newproc": pid, "orig": connection_record.info['pid']})
            connection_record.connection = connection_proxy.connection = None
            raise exc.DisconnectionError(
                "Connection record belongs to pid %s, "
                "attempting to check out in pid %s" %
                (connection_record.info['pid'], pid)
            )


if __name__ == '__main__':
    Base = declarative_base()
    pg_config = {'username': settings.PG_USER, 'password': settings.PG_PASSWORD, 'database': settings.PG_DBNAME,
                 'host': settings.DB_HOST}
    pg_dsn = "postgresql+psycopg2://{username}:{password}@{host}:5432/{database}".format(**pg_config)
    db_engine = create_engine(pg_dsn)
    add_engine_pidguard(db_engine)
    pg_meta = MetaData(bind=db_engine, schema="fintweet")

    class User(Base):
        __table__ = Table('user', pg_meta, autoload=True)
        tweets = relationship('Tweet')
        counts = relationship('UserCount')

    class UserCount(Base):
        __table__ = Table('user_count', pg_meta, autoload=True)

    class TweetCount(Base):
        __table__ = Table('tweet_count', pg_meta, autoload=True)

    class TweetMentions(Base):
        __table__ = Table('tweet_mentions', pg_meta, autoload=True)

    class TweetCashtags(Base):
        __table__ = Table('tweet_cashtags', pg_meta, autoload=True)

    class TweetHashtags(Base):
        __table__ = Table('tweet_hashtags', pg_meta, autoload=True)

    class TweetUrl(Base):
        __table__ = Table('tweet_url', pg_meta, autoload=True)

    class Tweet(Base):
        __table__ = Table('tweet', pg_meta, autoload=True)
        counts = relationship('TweetCount')
        ment_s = relationship('TweetMentions')
        cash_s = relationship('TweetCashtags')
        hash_s = relationship('TweetHashtags')
        url_s = relationship('TweetUrl')

    DstSession = sessionmaker(bind=db_engine, autoflush=False)
    # dstssn = DstSession()

    if True:  # settings.tweets:
        try:
            command = sys.argv[1]
        except IndexError:
            command = ''

        if command == 'nolocation':
            ISLOCATION = False
        else:
            ISLOCATION = True

        user_queue = multiprocessing.dummy.Queue()

        fname = 'word_list_3.xlsx'
        wb = load_workbook(fname)
        ws = wb.active
        ii = i = 2
        while True:
            if not ws.cell(row=i, column=1).value:
                break
            t1 = str(ws.cell(row=i, column=4).value).lower().strip(' ')
            t2 = str(ws.cell(row=i, column=5).value).lower().strip(' ')
            t1 = re.sub(' 00:00:00', '', t1)
            t2 = re.sub(' 00:00:00', '', t2)
            permno = str(ws.cell(row=i, column=1).value).lower().strip(' ')
            query = str(ws.cell(row=i, column=2).value).lower().strip(' '), \
                str(ws.cell(row=i, column=3).value).lower().strip(' '), \
                t1, t2, permno
            print(query)

            user_queue.put((query, i))
            i += 1

        pool = ThreadPool(len(settings.proxy_list))
        # pool = ThreadPool(8)
        lock = Lock()
        # Single process for testings
        # scrape_query(user_queue, '154.16.11.199:3199', lock, pg_dsn)
        # exit()
        pool.map(lambda x: (scrape_query(user_queue, x, lock, pg_dsn)), settings.proxy_list)
