import csv
from datetime import datetime
import json
import logging
import logging.config
from multiprocessing.dummy import Lock, Queue
from multiprocessing.dummy import Pool as ThreadPool
import os
import re
import requests
import sys
import time

from openpyxl import load_workbook
from sqlalchemy import create_engine, MetaData, Table
from sqlalchemy.orm import relationship, sessionmaker, scoped_session
from sqlalchemy.ext.declarative import declarative_base
import sqlalchemy.exc as exc
import sqlalchemy.event as event
import warnings

import settings


os.environ['MKL_NUM_THREADS'] = '1'
os.environ['OMP_NUM_THREADS'] = '1'
os.environ['MKL_DYNAMIC'] = 'FALSE'

pg_config = {
    'username': settings.PG_USER,
    'password': settings.PG_PASSWORD,
    'database': settings.PG_DBNAME,
    'host': settings.DB_HOST
}
pg_dsn = "postgresql+psycopg2://{username}:{password}@{host}/{database}".format(**pg_config)


def add_engine_pidguard(engine):
    @event.listens_for(engine, "connect")
    def connect(dbapi_connection, connection_record):
        connection_record.info['pid'] = os.getpid()

    @event.listens_for(engine, "checkout")
    def checkout(dbapi_connection, connection_record, connection_proxy):
        pid = os.getpid()
        if connection_record.info['pid'] != pid:
            # substitute log.debug() or similar here as desired
            warnings.warn("Parent process %(orig)s forked (%(newproc)s) with an open "
                          "database connection, "
                          "which is being discarded and recreated." % {
                              "newproc": pid,
                              "orig": connection_record.info['pid']
                          })
            connection_record.connection = connection_proxy.connection = None
            raise exc.DisconnectionError("Connection record belongs to pid %s, "
                                         "attempting to check out in pid %s" %
                                         (connection_record.info['pid'], pid))


Base = declarative_base()
db_engine = create_engine(
    pg_dsn,
    connect_args={"application_name": 'stocktwits_scraper:' + str(__name__)},
    pool_size=300,
    pool_recycle=600,
    max_overflow=0,
    encoding='utf-8'
    )
add_engine_pidguard(db_engine)
pg_meta = MetaData(bind=db_engine, schema="stocktwits")


class User(Base):
    __table__ = Table('user', pg_meta, autoload=True)
    ideas = relationship('Ideas')
    counts = relationship('User_Count')
    strategy = relationship('UserStrategy')


class User_Count(Base):
    __table__ = Table('user_count', pg_meta, autoload=True)


class UserStrategy(Base):
    __table__ = Table('user_strategy', pg_meta, autoload=True)


class Ideas(Base):
    __table__ = Table('ideas', pg_meta, autoload=True)
    counts = relationship('Ideas_Count')
    cash_s = relationship('IdeasCashtags')
    hash_s = relationship('IdeasHashtags')
    url_s = relationship('IdeasUrls')


class Ideas_Count(Base):
    __table__ = Table('ideas_count', pg_meta, autoload=True)


class IdeasCashtags(Base):
    __table__ = Table('idea_cashtags', pg_meta, autoload=True)


class IdeasHashtags(Base):
    __table__ = Table('idea_hashtags', pg_meta, autoload=True)


class IdeasUrls(Base):
    __table__ = Table('idea_urls', pg_meta, autoload=True)


class LoadingError(Exception):
    pass


emoticons_str = r"""
    (?:
        [:=;] # Eyes
        [oO\-]? # Nose (optional)
        [D\)\]\(\]/\\OpP] # Mouth
    )"""

regex_str = [
    emoticons_str,
    r'<[^>]+>',  # HTML tags
    r'(?:@[\w_]+)',  # @-mentions
    r"(?:\#[a-zA-Z0-9]\w+)",    # hash-tags
    r"(?:\B\$[A-Za-z][A-Za-z0-9]{0,4}\b)",    # cash-tags
    r'(?:http[s]?:\/\/(?:[a-z]|[0-9]|[$-_@.&amp;+]|[!*\(\),]|(?:%[0-9a-f][0-9a-f]))+)',  # URLs
    r'(?:(?:\d+,?)+(?:\.?\d+)?)',  # numbers
    r"(?:[a-z][a-z'\-_]+[a-z])",  # words with - and '
    r'(?:[\w_]+)',  # other words
    r'(?:\S)'  # anything else
]

tokens_re = re.compile(r'(' + '|'.join(regex_str) + ')', re.VERBOSE | re.IGNORECASE)
emoticon_re = re.compile(r'^' + emoticons_str + '$', re.VERBOSE | re.IGNORECASE)


def tokenize(s, lowercase=False):
    tokens = tokens_re.findall(s)
    if lowercase:
        tokens = [token if emoticon_re.search(token) else token.lower() for token in tokens]
    return tokens


class Page(object):
    def __init__(self, proxy=None):
        self.pr = {'http': proxy, 'https': proxy}
        self.timeout = 30
        self.ses = requests.Session()
        self.ses.headers = {
            'user-agent':
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36',
            # 'Connection': 'keep-alive',
            # 'Cache-Control': 'max-age=0',
            'Accept-Encoding':
            'gzip, deflate, br',
            'Origin':
            'https://stocktwits.com',
            # 'accept': 'application/json, text/javascript, */*; q=0.01',
            'Accept':
            '*/*',
            'Accept-Language':
            'en-US;q=0.6,en;q=0.4',
            'x-compress':
            None,
            # 'Upgrade-Insecure-Requests': '1',
            # 'x-requested-with': 'XMLHttpRequest',
            # 'x-twitter-active-user': 'yes',
            # 'host': 'stocktwits.com'
        }

    # def __del__(self):
    #     self.ses.close()
    #
    # def close(self):
    #     self.ses.close()

    def load(self, n, typ, url, params=None, headers=None, important=True):
        error_count = 0
        while True:
            try:
                if typ == 'get':
                    resp = self.ses.get(url, params=params, headers=headers, timeout=self.timeout)
                elif typ == 'put':
                    resp = self.ses.put(
                        url, data=json.dumps(params), headers=headers, timeout=self.timeout)
                elif typ == 'post':
                    resp = self.ses.post(
                        url, data=json.dumps(params), headers=headers, timeout=self.timeout)
                elif typ == 'delete':
                    resp = self.ses.delete(
                        url, data=json.dumps(params), headers=headers, timeout=self.timeout)

            except requests.exceptions.RequestException as e:
                error_count += 1
                # print('Loading error', error_count, pr, e)
                if error_count < 3:
                    logger.debug('%s Loading error %s %s %s %s', n, error_count, url, self.pr, e)
                    time.sleep(60)
                    continue
                logger.error('%s Error limit exceeded Loading error %s %s %s', n, url, self.pr, e)
                raise LoadingError

            if resp.status_code == requests.codes.ok:
                return resp.text
            elif resp.status_code == 429:
                logger.debug('%s Rate limit. Sleep 3 min', n)
                time.sleep(3 * 60)
                continue
            elif resp.status_code == 502 or resp.status_code == 504:
                logger.debug('%s wait 30 sec Error %s %s %s %s', n, resp.status_code, error_count, url, params)
                error_count += 1
                if error_count < 5:
                    time.sleep(30)
                else:
                    if important:
                        logger.error('%s Error limit exceeded HTTP error %s ', n, resp.status_code)
                        raise LoadingError
                    else:
                        return None
            elif resp.status_code == 503:
                # print('503 Error waiting 2 min', screen_name, proxy, url, resp.text, resp.status_code)
                error_count += 1
                if error_count > 5:
                    logger.error('%s Requestes 503 error %s %s', n, url, self.pr)
                    raise LoadingError
                logger.debug('%s Requestes 503 error %s %s %s', n, error_count, url, self.pr)
                time.sleep(120)
                continue
            else:
                error_count += 1
                # print(n, resp.text)
                logger.debug('%s HTTP Error %s %s %s %s', n, resp.status_code, error_count, url, params)
                if not important:
                    return None

                if error_count > 5:
                    logger.error('%s Error limit exceeded Requestes error %s %s %s', n, url, self.pr, resp.status_code)
                    raise LoadingError
                logger.debug('%s Error waiting 1 min %s %s', n, resp.status_code, url)
                time.sleep(60)
                continue


def get_symbols(s):
    s = s.upper()
    res = re.findall('(\$[A-Z]{1,6}([._][A-Z]{1,2})?)', s, re.M)
    if res:
        r = list(map(lambda x: x[0], res))
        r = list(map(lambda x: x, r))
        r = list(set(r))
        return r
    else:
        return []


def get_new_search(n, page, proxy, query):
    # page = Page(proxy)
    params = {}
    query = query.strip('$').upper()
    url = 'https://stocktwits.com/symbol/' + query
    r = page.load(n, 'get', url)
    params = {}
    params['filter'] = 'all'
    params['max'] = None
    u = 'https://api.stocktwits.com/api/2/streams/symbol/' + query + '.json'
    while True:
        r = page.load(n, 'get', u, params=params, headers={'referer': url})
        j = json.loads(r)
        if j['response']['status'] == 200:
            for message in j['messages']:
                # t = get_details(page, message, query)
                yield message
            if j['cursor']['more']:
                params['max'] = j['cursor']['max']
                continue


def login(n, page):
    url = 'https://stocktwits.com/api/login'
    h = {'content-type': 'application/json'}
    params = {}
    params['user_session'] = {}
    params['user_session']['login'] = settings.stocktwits_login
    params['user_session']['password'] = settings.stocktwits_password

    r = page.load(n, 'post', url, params=params, headers=h)
    try:
        j = json.loads(r)
        page.ses.headers['authorization'] = 'OAuth ' + j['token']
    except Exception:
        logger.error('%s Unable to login', n)
        logger.exception('message')
        sys.exit()
    logger.debug('%s Login success', n)


def insert_new_user(t, sess, watch_list=None):
    user = User(
        user_id=t['user']['id'],
        user_name=t['user']['name'],
        user_handle=t['user']['username'],
        date_joined=t['user']['join_date'],
        website=t['user']['website_url'],
        user_topmentioned=' '.join([f['symbol'] for f in t['user']['most_mentioned'][0]])[:255]
        if t['user'].get('most_mentioned', False) else None,
        verified='YES' if t['user']['official'] else 'NO',
        location=t['user']['location'][:255] if t['user'].get('location', False) else None
        )

    usr_st_dict = t['user']['trading_strategy']
    assets_frequently_traded = json.dumps(usr_st_dict['assets_frequently_traded'])
    user_strategy = UserStrategy(
        user_id=t['user']['id'],
        assets_frequently_traded=assets_frequently_traded,
        approach=usr_st_dict['approach'],
        holding_period=usr_st_dict['holding_period'],
        experience=usr_st_dict['experience'])

    user_count = User_Count(
        followers=t['user']['followers'],
        following=t['user']['following'],
        watchlist_count=t['user']['watchlist_stocks_count'],
        watchlist_stocks=watch_list,
        ideas=t['user']['ideas'])
    try:
        session = sess()
        user.counts.append(user_count)
        user.strategy.append(user_strategy)
        session.add(user_count)
        session.add(user_strategy)
        session.add(user)
        # session.flush()
        session.commit()
        logger.info('Inserted new user: %s %s', user.user_id, user.user_handle)
    except exc.IntegrityError as err:
        if re.match("(.*)Duplicate entry(.*)for key 'PRIMARY'(.*)", err.args[0]):
            logger.warning('%s ROLLBACK USER Duplicate entry %s', n, t['user']['username'])
            session.rollback()
    except Exception:
        logger.exception('message')
    finally:
        session.close()
        return user.user_id


def insert_new_idea(t, sess, reply_to=None):
    logger.debug('Inserting new idea')

    idea = Ideas(
        ideas_id=t['id'],
        user_id=t['user']['id'],
        datetime=t['created_at'],
        reply_to=reply_to if reply_to else None,
        text=t['body'],
        sentiment=t['entities']['sentiment']['basic'] if t['entities']['sentiment'] else None,
        permalink='https://stocktwits.com/' + t['user']['username'] + '/message/' + str(t['id'])
    )

    # if 'reshare_message' in t:
    #     pprint(t['reshare_message'])
    idea_count = Ideas_Count(
        replies=t['conversation']['replies'] if t.get('conversation', False) else None,
        likes=t['likes']['total'] if t.get('likes', False) else None,
        reshares=t['reshare_message']['reshared_count'] if t.get('reshare_message', False) else None
        )
    idea.counts.append(idea_count)

    tokens = tokenize(idea.text)
    cashtags = set([term.upper() for term in tokens if term.startswith('$') and len(term) > 2])
    if len(cashtags) > 0:
        for cashtag in cashtags:
            ctag = IdeasCashtags(ideas_id=idea.ideas_id, cashtag=cashtag[:64])
            idea.cash_s.append(ctag)
    hashtags = set([term.upper() for term in tokens if term.startswith('#') and len(term) > 2])
    if len(hashtags) > 0:
        for hashtag in hashtags:
            htag = IdeasHashtags(ideas_id=idea.ideas_id, hashtag=hashtag[:64])
            idea.hash_s.append(htag)

    if t.get('links', False):
        for u in t['links']:
            utag = IdeasUrls(ideas_id=idea.ideas_id, url=u['url'][:255])
            idea.url_s.append(utag)

    try:
        session = sess()
        session.add(idea_count)
        session.add(idea)
        session.commit()
        if reply_to:
            logger.info('Inserted new Idea: %s as reply to %s', idea.ideas_id, idea.reply_to)
        else:
            logger.info('Inserted new Idea: %s', idea.ideas_id)
    except exc.IntegrityError as err:
        if re.match("(.*)Duplicate entry(.*)for key 'PRIMARY'(.*)", err.args[0]):
            logger.warning('%s ROLLBACK IDEAS DUPLICATE ENTRY', n)
            session.rollback()
    except Exception:
        logger.exception('message')
        print(idea)
    finally:
        session.close()


def get_tweets(n, dateto, permno, proxy, query, lock):
    ScopedSession = scoped_session(sessionmaker(bind=db_engine))
    query = query.strip('$').upper()
    count = 0
    page = Page(proxy)
    login(n, page)
    for t in get_new_search(n, page, proxy, query):
        count_repl = 0
        t1 = datetime.strptime(t['created_at'], '%Y-%m-%dT%H:%M:%SZ')
        if dateto > t1:
            logger.debug('dateto > t1')
            break
        session = ScopedSession()
        idea = session.query(Ideas).filter_by(ideas_id=t['id']).first()
        if idea:
            logger.debug('Idea id [%s] exists in db', t['id'])
            continue

        user = session.query(User).filter_by(user_id=t['user']['id']).first()
        if not user:
            try:
                if settings.ISUSERPROFILE:
                    url = 'https://stocktwits.com/api/user_info_stream/' + t['user']['username'] + '?limit=15'
                    r = page.load(n, 'get', url, headers={'referer': 'https://stocktwits.com/' + t['user']['username']},
                                  important=False)
                    if r:
                        j = json.loads(r)
                        t['user'] = j['user']
                if settings.ISWATCHLIST:
                    url = 'https://api.stocktwits.com/api/2/watchlists/user/' + str(t['user']['id']) + '.json'
                    r = page.load(n, 'get', url, headers={'referer': 'https://stocktwits.com/' + t['user']['username']},
                                  important=False)
                    if r:
                        j = json.loads(r)
                        watch_list = ' '.join([f['symbol'] for f in j['watchlist']['symbols']])[:500]
                else:
                    watch_list = None
                insert_new_user(t, ScopedSession, watch_list)
            except Exception:
                logger.exception('message')
        if t.get('conversation', False):
            reply_to = t['conversation']['in_reply_to_message_id']
        else:
            reply_to = None
        idea_id = insert_new_idea(t, ScopedSession, reply_to)

        if t.get('conversation', False) and settings.GET_REPLYS:
            count_repl = 0
            url = 'https://api.stocktwits.com/api/2/messages/' + str(t['id']) + '/conversation.json?max='
            flag_conv = False
            while True:
                r = page.load(n, 'get', url, important=False)
                if not r:
                    logger.error('%s Error loading conversation', n)
                    break    # , headers={'referer': 'https://stocktwits.com/' + t['user']['username']})
                j = json.loads(r)
                if flag_conv:
                    mess = j
                else:
                    mess = j['children']
                for child in mess['messages']:
                    count_repl += 1
                    reply = session.query(Ideas).filter_by(ideas_id=child['id']).first()
                    user = session.query(User).filter_by(user_id=child['user']['id']).first()
                    session.close()
                    if not user:
                        try:
                            if settings.ISUSERPROFILE:
                                url = 'https://stocktwits.com/api/user_info_stream/' + child['user']['username'] \
                                    + '?limit=15'
                                r = page.load(n, 'get', url, headers={'referer': 'https://stocktwits.com/'
                                              + child['user']['username']}, important=False)
                                if r:
                                    j = json.loads(r)
                                    child['user'] = j['user']
                            if settings.ISWATCHLIST:
                                url = 'https://api.stocktwits.com/api/2/watchlists/user/' + str(child['user']['id']) \
                                    + '.json'
                                r = page.load(n, 'get', url, headers={'referer': 'https://stocktwits.com/'
                                              + child['user']['username']}, important=False)
                                if r:
                                    j = json.loads(r)
                                    watch_list = ' '.join([f['symbol'] for f in j['watchlist']['symbols']])[:500]
                            else:
                                watch_list = None
                            insert_new_user(child, ScopedSession, watch_list)
                        except Exception:
                            logger.exception('message')
                    if not reply:
                        reply_id = insert_new_idea(child, ScopedSession, reply_to=t['id'])
                        count += 1

                if mess['cursor']['more']:
                    url = 'https://api.stocktwits.com/api/2/messages/' + str(t['id']) + '/children.json?since=' \
                        + str(mess['cursor']['since'])
                    flag_conv = True
                    continue
                break
        ScopedSession.remove()
        count += 1

    if count:
        fname = 'report.csv'
        fdnames = [
            'time',
            'cashtag',
            'permno',
            'number',
        ]
        data = {}
        data['time'] = time.strftime('%Y-%m-%d %H:%M:%S')
        data['cashtag'] = query

        data['permno'] = permno
        data['number'] = count

        if os.path.isfile(fname):

            with open(fname, 'a', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, lineterminator='\n', fieldnames=fdnames,
                                        dialect='excel', quotechar='"', quoting=csv.QUOTE_ALL)
                writer.writerow(data)
        else:
            with open(fname, 'w') as f:
                writer = csv.DictWriter(f, lineterminator='\n', fieldnames=fdnames,
                                        dialect='excel', quotechar='"', quoting=csv.QUOTE_ALL)
                writer.writeheader()
                writer.writerow(data)
    else:
        pass
    return count


def scrape(n, user_queue, proxy, lock):
    # db_engine = create_engine(pg_dsn, pool_size=1)
    # add_engine_pidguard(db_engine)
    # Session = sessionmaker(bind=db_engine)
    # session = Session()

    while not user_queue.empty():
        dateto, permno, query, i = user_queue.get()
        n = i
        logger.info('%2s START %2s %s %s', n, i, proxy, query)
        try:
            res = get_tweets(n, dateto, permno, proxy, query, lock)
        except LoadingError:
            logger.error('%2s LoadingError except', n)
            return
        except Exception:
            logger.exception('message')
            raise
        if not res:
            logger.error('%2s SCRAP_USER Error in %s %s', n, query, i)
            # with open('error_list.txt', 'a') as f:
            #     f.write(query[0] + '\n')
        else:
            logger.error('%2s ENDED %2s %s %s %s', n, i, proxy, query, res)

        time.sleep(1)


def create_logger():
    log_file = 'stocktwits_scrapper_' + str(datetime.now().strftime('%Y-%m-%d')) + '.log'
    logging.config.fileConfig('log.ini', defaults={'logfilename': log_file})
    return logging.getLogger(__name__)


logger = create_logger()

if __name__ == '__main__':
    # multiprocessing.set_start_method('forkserver')

    # Session = sessionmaker(bind=db_engine)
    # session = Session()

    user_queue = Queue()
    lock = Lock()

    fname = 'symbol_list.xlsx'
    try:
        wb = load_workbook(fname)
        ws = wb.active
    except Exception:
        logger.error('Could not open file %s. Exitting...', fname)
        # logger.exception('message')
        exit()
    i = 2
    while True:
        if not ws.cell(row=i, column=1).value:
            break
        permno = str(ws.cell(row=i, column=1).value).lower().strip(' ')
        query = str(ws.cell(row=i, column=2).value).lower().strip(' ')
        dd = ws.cell(row=i, column=3).value
        dateto = dd if isinstance(dd, datetime) else datetime.strptime(dd, '%d/%m/%Y')

        # print(query)
        user_queue.put((dateto, permno, query, i - 1))
        i += 1

    np = len(settings.proxy_list) if len(settings.proxy_list) > i - 2 else i - 2
    pp = []
    proxy_list = []
    n = 0
    for i, proxy in enumerate(settings.proxy_list):
        n += 1
        proxy_list.append({'proxy': proxy, 'num': n})
    try:
        # Single process for testing
        # scrape(0, user_queue, proxy_list[0], lock)
        # exit()
        pool = ThreadPool(len(settings.proxy_list))
        while True:
            pool.map(lambda x: (scrape(x['num'], user_queue, x['proxy'], lock)), proxy_list)
            pool.close()
            pool.join()
            if not user_queue.empty():
                continue
    except KeyboardInterrupt:
        logger.warning('Interrupted by keyboard.')
    except Exception:
        logger.exception('message')
    finally:
        db_engine.dispose()
        logger.info('db_engine disposed')
        logger.info('Exiting...')
        exit()
