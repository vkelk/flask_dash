import os.path
from multiprocessing.dummy import Pool as ThreadPool, Lock
from multiprocessing import Process
import csv
import datetime
import json
import multiprocessing.dummy
import re
import requests
import sys
import time
from dateutil import parser as dateparser
from openpyxl import load_workbook
from pyquery import PyQuery
from datetime import datetime
import settings
from sqlalchemy import *
from sqlalchemy import event, exc
from sqlalchemy.orm import sessionmaker, scoped_session
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import relationship
import warnings
import os
import sqlalchemy.exc
from pprint import pprint
import random
import tweet_api

ISUSERPROFILE = True
IS_PROFILE_SEARCH=True
time_wait = 0
flag1 = False

emoticons_str = r"""
    (?:
        [:=;] # Eyes
        [oO\-]? # Nose (optional)
        [D\)\]\(\]/\\OpP] # Mouth
    )"""
regex_str = [
    emoticons_str,
    r'<[^>]+>',  # HTML tags
    r'(?:@[\w_]+)',  # @-mentions
    r"(?:\#+[\w_]+[\w\'_\-]*[\w_]+)",  # hash-tags
    r"(?:\$+[a-zA-Z]+[\w\'_\-]*[\w_]+)",  # cash-tags
    r'http[s]?://(?:[a-z]|[0-9]|[$-_@.&amp;+]|[!*\(\),]|(?:%[0-9a-f][0-9a-f]))+',  # URLs

    r'(?:(?:\d+,?)+(?:\.?\d+)?)',  # numbers
    r"(?:[a-z][a-z'\-_]+[a-z])",  # words with - and '
    r'(?:[\w_]+)',  # other words
    r'(?:\S)'  # anything else
]
tokens_re = re.compile(r'(' + '|'.join(regex_str) + ')', re.VERBOSE | re.IGNORECASE)
emoticon_re = re.compile(r'^' + emoticons_str + '$', re.VERBOSE | re.IGNORECASE)


class LoadingError(Exception):
    pass


class Twit:
    def __init__(self):
        pass

    def json(self):
        return {'date': self.unixtime, 'text': self.text, 'screen_name': self.screen_name, 'user_name': self.user_name,
                'user_id': self.user_id,
                'id': self.id, 'retweets_count': self.retweets_count, 'favorites_count': self.favorites_count,
                'permalink': self.permalink, 'urls': self.urls, 'mentions': self.mentions,
                'hashtags': self.hashtags, 'is_retweet': self.is_retweet, 'symbols': self.symbols,
                'is_protected': self.is_protected}

    def __repr__(self):
        return self.text





def get_symbols(s):
    s = s.upper()
    res = re.findall('(\$[A-Z]{1,6}([._][A-Z]{1,2})?)', s, re.M)
    if res:
        r = list(map(lambda x: x[0], res))
        r = list(set(r))
        return r
    else:
        return []


def scra(query, i, proxy, lock, session):
    twitter_scraper=tweet_api.TweetScraper(proxy, IS_PROFILE_SEARCH=IS_PROFILE_SEARCH, logname='awam')
    def tokenize(s):
        return tokens_re.findall(s)

    def preprocess(s, lowercase=False):
        tokens = tokenize(s)
        if lowercase:
            tokens = [token if emoticon_re.search(token) else token.lower() for token in tokens]
        return tokens

    permno = query[4]
    fieldnames = ["QueryStartDate", "QueryEndDate", "Query", "DateOfActivity", "UserScreenName", "Keyword",
                  "Location", "Website", "DateJoined", "IsMention", "UserID", "TimeOfActivity", "Hashtags",
                  "Re_tweet", "NumberOfReplies", "NumberOfRe_tweets", "NumberOfFavorites", "Tweet", "tweet_id",
                  "tweet_url", "is_verified", "Urls", "UserFollowersCount", "UserFollowingCount", "UserTweetsCount",
                  "LikesCount", "CashtagSymbols", "user_location", "permno"]

    count = 0

    ttm = time.time()
    tweet_list = []

    # Receiving random credentials from the settings file
    credential = settings.twitter_logins[random.randrange(0, len(settings.twitter_logins))]
    login = credential['login']
    password = credential['password']

    for t in twitter_scraper.get_new_search( query, login, password):

        data = {}
        data['permno'] = permno
        data['user_location'] = t.user_location
        data['LikesCount'] = t.likes
        data['Website'] = t.website
        data['QueryStartDate'] = query[2]
        data['QueryEndDate'] = query[3]
        data['Query'] = query[1]
        data['Keyword'] = query[1]
        data['TimeOfActivity'] = time.strftime('%H:%M:%S', time.localtime(t.unixtime))
        data['DateOfActivity'] = time.strftime('%d/%m/%Y', time.localtime(t.unixtime))
        data['tweet_id'] = t.id
        data['tweet_url'] = t.permalink
        data['UserID'] = t.user_id
        data['UserScreenName'] = t.screen_name
        data['UserName'] = t.user_name
        data['TimeZone'] = t.time_zone
        data['UserTweetsCount'] = t.user_tweet_count
        data['UserFollowersCount'] = t.user_followers_count
        data['UserFollowingCount'] = t.user_following_count
        data['NumberOfFavorites'] = t.favorites_count
        data['NumberOfRe_tweets'] = t.retweets_count
        data['NumberOfReplies'] = t.replyes_count
        data['Re_tweet'] = t.is_retweet
        data['is_verified'] = t.is_verified
        # data['isProtected'] = t.is_protected
        data['isReply'] = t.is_reply
        data['ReplyTweetID'] = t.data_conversation_id
        data['ReplyUserId'] = t.is_reply_id
        # data['ReplyScreenName'] = t.is_reply_screen_name
        # data['Lang'] = t.lang
        data['Hashtags'] = ' '.join(t.hashtags)
        data['Urls'] = ', '.join(t.urls)
        data['CashtagSymbols'] = '\n'.join(t.symbols)
        # data['Mentions_id'] = '\n'.join(t.mentions_id)
        data['IsMention'] = ' '.join(t.mentions_name)
        data['Location'] = t.location_name
        # data['Location ID'] = t.location_id
        data['DateJoined'] = dateparser.parse(t.user_created) if t.user_created else None
        data['Tweet'] = t.text

        if t.utc_offset:
            if t.utc_offset / 3600 >= 0:
                data['TimeZoneUTC'] = 'UTC+' + str(int(t.utc_offset / 3600))
            else:
                data['TimeZoneUTC'] = 'UTC' + str(int(t.utc_offset / 3600))
        else:
            data['TimeZoneUTC'] = None


        tokens = preprocess(t.text)
        cashtags = [term for term in tokens if term.startswith('$') and len(term) > 1]
        hashtags = [term for term in tokens if term.startswith('#') and len(term) > 1]
        mentions = [term for term in tokens if term.startswith('@') and len(term) > 1]
        urls = [term for term in tokens if term.startswith('http') and len(term) > 4]

        tweet_list.append(data)
        # pprint(data)

        print(query, count, t.date)
        # print(data)
        if ISUSERPROFILE:
            pass
            # dd = re.findall('\w+ (\w+) (\d+) \d+:\d+:\d+ \+\d+ (\d+)', data['DateJoined'])
            # try:
            #     date_joined = datetime.strptime(' '.join(dd[0]), '%b %d %Y')
            # except IndexError:
            #     print(data['DateJoined'], dd)
            #     exit()
        else:
            date_joined = None
            t.user_listed_count = None
            t.username = t.user_name
            t.user_timezone = None
            t.user_description = None
            data['Website'] = None
            data['is_verified'] = None

        if session.query(Tweet).filter_by(tweet_id=data['tweet_id']).first():
            continue

        user = session.query(User).filter_by(user_id=data['UserID']).first()
        if not user:
            user_count = UserCount(follower=data['UserFollowersCount'],
                                   following=data['UserFollowingCount'],
                                   tweets=data['UserTweetsCount'],
                                   likes=data['NumberOfFavorites'],
                                   lists=t.user_listed_count)
            session.add(user_count)
            user = User(user_id=data['UserID'],
                        twitter_handle=data['UserScreenName'][:120],
                        user_name=data['UserName'][:120],
                        location=data['user_location'][:255] if data['user_location'] else None,
                        date_joined=data['DateJoined'],
                        timezone=data['TimeZoneUTC'],
                        website=data['Website'][:255] if data['Website'] else None,
                        user_intro=t.user_description[:255] if t.user_description else None,
                        verified=data['is_verified'])
            user.counts.append(user_count)
            session.add(user)
            try:
                session.commit()
            except sqlalchemy.exc.IntegrityError as err:
                if re.search("duplicate key value violates unique constraint", err.args[0]):
                    print('ROLLBACK USER')
                    session.rollback()
            except Exception as e:
                print(e)
                raise

        twit = session.query(Tweet).filter_by(tweet_id=data['tweet_id']).first()
        if not twit:
            twit = Tweet(tweet_id=data['tweet_id'],
                         date=datetime.strptime(data['DateOfActivity'], '%d/%m/%Y'),
                         time=data['TimeOfActivity'],
                         timezone=data['TimeZoneUTC'][:10] if data['TimeZoneUTC'] else None,
                         retweet_status=data['Re_tweet'],
                         text=data['Tweet'],
                         reply_to=t.data_conversation_id if t.is_reply else None,
                         location=data['Location'][:255] if data['Location'] else None,
                         permalink=data['tweet_url'] if data['tweet_url'] else None,
                         emoticon=','.join(t.emoji) if t.emoji else None)
            tweet_count = TweetCount(reply=data['NumberOfReplies'],
                                     favorite=data['NumberOfFavorites'],
                                     retweet=t.retweets_count)

        if not session.query(TweetHashtags).filter_by(tweet_id=data['tweet_id']).first():
            for hash_s in hashtags:
                tweet_hashtag = TweetHashtags(hashtags=hash_s[:45])
                twit.hash_s.append(tweet_hashtag)
                session.add(tweet_hashtag)

        if not session.query(TweetUrl).filter_by(tweet_id=data['tweet_id']).first():
            for url_s in urls:
                tweet_url = TweetUrl(url=url_s[:255])
                twit.url_s.append(tweet_url)
                session.add(tweet_url)

        if not session.query(TweetCashtags).filter_by(tweet_id=data['tweet_id']).first():
            for cash_s in cashtags:
                tweet_cashtag = TweetCashtags(cashtags=cash_s[:45])
                twit.cash_s.append(tweet_cashtag)
                session.add(tweet_cashtag)

        if not session.query(TweetMentions).filter_by(tweet_id=data['tweet_id']).first():
            for ment_s in t.ment_s:
                tweet_mentions = TweetMentions(mentions=ment_s[0][:45], user_id=ment_s[1])
                twit.ment_s.append(tweet_mentions)
                session.add(tweet_mentions)
        user.tweets.append(twit)

        if not session.query(TweetCount).filter_by(tweet_id=int(data['tweet_id'])).first():
            twit.counts.append(tweet_count)
            session.add(tweet_count)
        else:
            i=1
        session.add(twit)
        try:
            session.commit()
        except sqlalchemy.exc.IntegrityError as err:
            if re.search('duplicate key value violates unique constraint', err.args[0]):
                print('ROLLBACK common')
                session.rollback()
        except Exception as e:
            print(e)
            raise
        count += 1

    lock.acquire()
    if count > 0:

        with open('report.csv', 'a') as f:
            data = {}
            fdnames = ['time', 'query_name', 'number', ]
            writer = csv.DictWriter(f, lineterminator='\n', fieldnames=fdnames, dialect='excel', quotechar='"',
                                    quoting=csv.QUOTE_ALL)
            data['time'] = time.strftime('%Y-%m-%d %H:%M:%S')
            data['query_name'] = query[1]
            data['number'] = count
            # data['date'] = t.user_created

            writer.writerow(data)
        lock.release()
        return count
    else:
        with open('error.csv', 'a') as f:
            data = {}
            fdnames = ['time', 'query_name', 'number', ]
            writer = csv.DictWriter(f, lineterminator='\n', fieldnames=fdnames, dialect='excel', quotechar='"',
                                    quoting=csv.QUOTE_ALL)
            data['time'] = time.strftime('%Y-%m-%d %H:%M:%S')
            data['query_name'] = query[1]
            data['number'] = count
            writer.writerow(data)
        lock.release()
        return False


def scrape_query(user_queue, proxy, lock, pg_dsn):
    db_engine = create_engine(pg_dsn, pool_size=1)
    add_engine_pidguard(db_engine)
    DstSession = sessionmaker(bind=db_engine, autoflush=False)
    session = DstSession()

    active = True
    while not user_queue.empty():
        query, i = user_queue.get()

        print('START', i, proxy, query)
        # TODO filter:nativeretweets
        try:
            res = scra(query, i, proxy, lock, session=session)
        except LoadingError:
            print('LoadingError except')
            return False
        if not res:
            print('     SCRAP_USER Error in', query, i)
            with open('error_list.txt', 'a') as f:
                f.write(query[0] + '\n')
        else:
            print(datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ' ENDED', i, proxy, query, res)
    return True


def get_up(fname, proxy):
    n = 0
    fieldnames = ["QueryStartDate", "QueryEndDate", "Query", "DateOfActivity", "UserScreenName", "Keyword",
                  "Location", "Website", "DateJoined", "IsMention", "UserID", "TimeOfActivity", "Hashtags",
                  "Re_tweet", "NumberOfReplies", "NumberOfRe_tweets", "NumberOfFavorites", "Tweet", "tweet_id",
                  "tweet_url", "is_verified", "Urls", "UserFollowersCount", "UserFollowingCount", "UserTweetsCount",
                  "LikesCount", "CashtagSymbols", "user_location", "permno"]
    # print('Processing file {} ...'.format(fname))
    with open(fname, 'r', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        recordlist = []
        count = 0
        for row in reader:
            recordlist.append(row)
            count += 1
    # print('   Read {} rows'.format(count))

    page = Page(proxy)
    new_list = []
    i = 0
    user_profiles = {}
    tt = time.time()
    for row in recordlist:
        tweet = Tweet()
        try:
            if user_profiles.get(row['UserScreenName'], None):
                r = Tweet()
                rt = user_profiles[row['UserScreenName']]
                r.user_location = rt['user_location']
                r.user_tweet_count = rt['UserTweetsCount']
                r.user_following_count = rt['UserFollowingCount']
                r.user_followers_count = rt['UserFollowersCount']
                r.likes = rt['LikesCount']
                r.user_created = rt['DateJoined']
                r.website = rt['Website']
                r.is_verified = rt['is_verified']

            else:
                try:
                    r = get_user_profile(row['UserScreenName'], page, tweet)
                except LoadingError:
                    print('Skip user', row['UserScreenName'])
                    r = None

            if r:
                tweet = r

                # print(tweet.likes,tweet.user_tweet_count,tweet.user_following_count,tweet.user_followers_count,
                # tweet.user_created,tweet.is_verified,tweet.website,tweet.user_location)
                rt = {}
                rt['user_location'] = row['user_location'] = tweet.user_location
                rt['UserTweetsCount'] = row['UserTweetsCount'] = tweet.user_tweet_count
                rt['UserFollowingCount'] = row['UserFollowingCount'] = tweet.user_following_count
                rt['UserFollowersCount'] = row['UserFollowersCount'] = tweet.user_followers_count
                rt['LikesCount'] = row['LikesCount'] = tweet.likes
                rt['DateJoined'] = row['DateJoined'] = tweet.user_created
                rt['Website'] = row['Website'] = tweet.website
                rt['is_verified'] = row['is_verified'] = tweet.is_verified
                user_profiles[row['UserScreenName']] = rt
            new_list.append(row)
            i += 1

            k = 50
            if i % k == 0:
                print('  {} Loaded {} user profiles '.format(fname, i))
                tt = time.time()
        except KeyError as e:
            print('File {} has the wrong content,skipped'.format(fname))
            break
    else:
        with open(fname, 'w', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, extrasaction='ignore', restval='', lineterminator='\n',
                                    fieldnames=fieldnames,
                                    dialect='excel', quotechar='"',
                                    quoting=csv.QUOTE_ALL)
            writer.writeheader()
            writer.writerows(new_list)
        # print('Write {} rows'.format(len(new_list)))
        n = len(new_list)

    return n


def get_pup(user_queue, proxy):
    while not user_queue.empty():
        fname, i = user_queue.get()

        print('START', i, proxy, fname)

        res = get_up(fname, proxy)

        print('ENDED', i, proxy, fname, res)
    return True


def add_engine_pidguard(engine):
    @event.listens_for(engine, "connect")
    def connect(dbapi_connection, connection_record):
        connection_record.info['pid'] = os.getpid()

    @event.listens_for(engine, "checkout")
    def checkout(dbapi_connection, connection_record, connection_proxy):
        pid = os.getpid()
        if connection_record.info['pid'] != pid:
            # substitute log.debug() or similar here as desired
            warnings.warn(
                "Parent process %(orig)s forked (%(newproc)s) with an open "
                "database connection, "
                "which is being discarded and recreated." %
                {"newproc": pid, "orig": connection_record.info['pid']})
            connection_record.connection = connection_proxy.connection = None
            raise exc.DisconnectionError(
                "Connection record belongs to pid %s, "
                "attempting to check out in pid %s" %
                (connection_record.info['pid'], pid)
            )


Base = declarative_base()
pg_config = {'username': settings.PG_USER, 'password': settings.PG_PASSWORD, 'database': settings.PG_DBNAME,
                'host': settings.DB_HOST}
pg_dsn = "postgresql+psycopg2://{username}:{password}@{host}:5432/{database}".format(**pg_config)
db_engine = create_engine(pg_dsn)
add_engine_pidguard(db_engine)
pg_meta = MetaData(bind=db_engine, schema="fintweet")

class User(Base):
    __table__ = Table('user', pg_meta, autoload=True)
    tweets = relationship('Tweet')
    counts = relationship('UserCount')


class UserCount(Base):
    __table__ = Table('user_count', pg_meta, autoload=True)


class TweetCount(Base):
    __table__ = Table('tweet_count', pg_meta, autoload=True)


class TweetMentions(Base):
    __table__ = Table('tweet_mentions', pg_meta, autoload=True)


class TweetCashtags(Base):
    __table__ = Table('tweet_cashtags', pg_meta, autoload=True)


class TweetHashtags(Base):
    __table__ = Table('tweet_hashtags', pg_meta, autoload=True)


class TweetUrl(Base):
    __table__ = Table('tweet_url', pg_meta, autoload=True)


class Tweet(Base):
    __table__ = Table('tweet', pg_meta, autoload=True)
    counts = relationship('TweetCount')
    ment_s = relationship('TweetMentions')
    cash_s = relationship('TweetCashtags')
    hash_s = relationship('TweetHashtags')
    url_s = relationship('TweetUrl')
    
DstSession = sessionmaker(bind=db_engine, autoflush=False)


if __name__ == '__main__':
    dstssn = DstSession()

    if True:  # settings.tweets:
        try:
            command = sys.argv[1]
        except IndexError:
            command = ''

        if command == 'nolocation':
            ISLOCATION = False
        else:
            ISLOCATION = True

        user_queue = multiprocessing.dummy.Queue()

        fname = 'input_profile.xlsx'
        wb = load_workbook(fname)
        ws = wb.active
        ii = i = 2
        while True:
            if not ws.cell(row=i, column=1).value:
                break

            permno = str(ws.cell(row=i, column=1).value).lower().strip(' ')
            query = str(ws.cell(row=i, column=1).value).lower().strip(' '), \
                    str(ws.cell(row=i, column=2).value).lower().strip(' ') ,0,0,None
            print(query)

            user_queue.put((query, i))
            i += 1

        pool = ThreadPool(len(settings.proxy_list))
        lock = Lock()
        pool.map(lambda x: (scrape_query(user_queue, x, lock, pg_dsn)), settings.proxy_list)
